import io
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.generics import RetrieveAPIView
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.request import Request
from rest_framework.response import Response
from app_deploy.general.utilitarios import getMonthName
from app_deploy.general.excel_utils import create_excel_from_json
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import xml.etree.ElementTree as ET

# Create your views here.

from .transporte import *

class TransporteVigenteView(RetrieveAPIView):

    def get(self, request: Request):        
        data = S09TranspVigente()

        return Response(
            data={"message": "Lista de transportes vigentes", "content": data},
            status=status.HTTP_200_OK,
        )

class TranspxAnioView(RetrieveAPIView):

    def get(self, request: Request):        
        c_anio = request.query_params.get("anio")
        data = S09TranspxAnio(c_anio)

        return Response(
            data={"message": "Vehiculos autorizados", "content": data},
            status=status.HTTP_200_OK,
        )

class ComparaTranspxAnioView(RetrieveAPIView):

    def get(self, request: Request):        
        m_dia = request.query_params.get("dia")
        m_mes = request.query_params.get("mes")
        c_anio01 = request.query_params.get("anio01")
        c_anio02 = request.query_params.get("anio02")
        opcion = request.query_params.get("opcion", 1)

        opcion = int(opcion)

        if opcion == 1 and (m_dia ==None or m_mes == None or c_anio01 == None or c_anio02 == None):
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )
            
        data = S09ComparaTranspxAnio(m_dia, m_mes, c_anio01, c_anio02, opcion)        

        return Response(
            data={"message": "Comparacion de vehiculos autorizados", "content": data},
            status=status.HTTP_200_OK,
        )
    
class TranspxAnioyMesView(RetrieveAPIView):

    def get(self, request: Request):        
        c_anio = request.query_params.get("anio")
        data = S09TranspxAnioyMes(c_anio)

        return Response(
            data={"message": "Vehiculos autorizados por mes", "content": data},
            status=status.HTTP_200_OK,
        )

class InfraccionesTransportexAnioView(RetrieveAPIView):
    
    def get(self, request: Request):        
        c_anio = request.query_params.get("anio")    
        
        data = S05InfraccionesTransportexAnio(c_anio)

        return Response(
            data={"message": "Infracciones de transporte", "content": data},
            status=status.HTTP_200_OK,
        )
        
class ComparaInfraccTransportexAnioView(RetrieveAPIView):

    def get(self, request: Request):        
        m_dia = request.query_params.get("dia")
        m_mes = request.query_params.get("mes")
        c_anio01 = request.query_params.get("anio01")
        c_anio02 = request.query_params.get("anio02")

        if m_dia ==None or m_mes == None or c_anio01 == None or c_anio02 == None:
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )
            
        data = S05ComparaInfraccTransportexAnio(m_dia, m_mes, c_anio01, c_anio02)

        return Response(
            data={"message": "Comparacion de infracciones de transporte", "content": data},
            status=status.HTTP_200_OK,
        ) 

class TranspAntigVehicView(RetrieveAPIView):

    def get(self, request: Request):        
        
        
        data = S09TranspAntigVehic()

        return Response(
            data={"message": "Vehiculos antiguos", "content": data},
            status=status.HTTP_200_OK,
        )
    
class OcurrenciasxAnioView(RetrieveAPIView):
    
        def get(self, request: Request):        
            c_opcion = request.query_params.get("opcion")
            c_anio = request.query_params.get("anio")
    
            if c_opcion == None or c_anio == None:
                return Response(
                    data={"message": "Faltan parametros"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
                
            data = S09OcurrenciasxAnio(c_opcion, c_anio)
    
            return Response(
                data={"message": "Ocurrencias por año", "content": data},
                status=status.HTTP_200_OK,
            )
        
class MontosPapeletaTransitoView(RetrieveAPIView):
    
        def get(self, request: Request):        
            anio = request.query_params.get("anio")
            tipo_infraccion = request.query_params.get("tipo")
            mes = request.query_params.get("mes")
    
            if anio == None or tipo_infraccion == None:
                return Response(
                    data={"message": "Faltan parametros"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
                
            data = S09MontosPapeletaTransito(anio, tipo_infraccion, mes)
    
            return Response(
                data={"message": "Montos de papeletas de transito", "content": data},
                status=status.HTTP_200_OK,
            )
        
class ComparaMontosPapeletaTransitoView(RetrieveAPIView):
    
        def get(self, request: Request):        
            m_dia = request.query_params.get("dia")
            m_mes = request.query_params.get("mes")
            c_anio01 = request.query_params.get("anio01")
            c_anio02 = request.query_params.get("anio02")
    
            if m_dia ==None or m_mes == None or c_anio01 == None or c_anio02 == None:
                return Response(
                    data={"message": "Faltan parametros"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
                
            data = S05ComparaMontosPapeletaTransito(m_dia, m_mes, c_anio01, c_anio02)
    
            return Response(
                data={"message": "Comparacion de montos de papeletas de transito", "content": data},
                status=status.HTTP_200_OK,
            )


class S42CapacitacionController(APIView):
    permission_classes = (IsAuthenticated,)

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get(self, request: Request):
        opcion = request.query_params.get('opcion')
        valor1 = request.query_params.get('valor1')
        valor2 = request.query_params.get('valor2')

        if not opcion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        capacitacion = S42SelectCapacitacion(opcion, valor1, valor2)

        return Response(data={
            "message": "Capacitación obtenida correctamente",
            "content": capacitacion
        }, status=status.HTTP_200_OK)     

    def post(self, request: Request):
        fecha = request.data.get('fecha')
        tema = request.data.get('tema')
        modalidad = request.data.get('modalidad')
        capacitador = request.data.get('capacitador')
        empresas = request.data.get('empresas')
        lugar = request.data.get('lugar')
        cantidad = request.data.get('cantidad')
        observacion = request.data.get('observacion')
        usuario = request.user.username

        if not fecha or not tema or not modalidad or not capacitador or not empresas or not lugar or not cantidad or not observacion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        capacitacion = S42InsertarCapacitacion(fecha, tema, modalidad, capacitador, empresas, lugar, cantidad, observacion, usuario)

        return Response(data={
            "message": "Capacitación insertada correctamente",
            "content": capacitacion
        }, status=status.HTTP_201_CREATED)  

    def put(self, request: Request, capacitacion: int):
        fecha = request.data.get('fecha')
        tema = request.data.get('tema')
        modalidad = request.data.get('modalidad')
        capacitador = request.data.get('capacitador')
        empresas = request.data.get('empresas')
        lugar = request.data.get('lugar')
        cantidad = request.data.get('cantidad')
        observacion = request.data.get('observacion')
        usuario = request.user.username

        if not fecha or not tema or not modalidad or not capacitador or not empresas or not lugar or not cantidad or not observacion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        S42UpdateCapacitacion(capacitacion, fecha, tema, modalidad, capacitador, empresas, lugar, cantidad, observacion, usuario)

        return Response(data={
            "message": "Capacitación actualizada correctamente",
            "content": None
        }, status=status.HTTP_200_OK)


    def delete(self, request: Request, capacitacion: int):
        usuario = request.user.username

        S42DeleteCapacitacion(capacitacion, usuario)

        return Response(data={
            "message": "Capacitación eliminada correctamente",
            "content": None
        }, status=status.HTTP_200_OK)

class S42CapacitacionObservacionController(APIView):
    permission_classes = (IsAuthenticated,)

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get(self, request: Request):
        opcion = request.query_params.get('opcion')
        valor1 = request.query_params.get('valor1')
        valor2 = request.query_params.get('valor2')

        if not opcion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        observacion = S42SelectCapacitacionObservacion(opcion, valor1, valor2)

        return Response(data={
            "message": "Observacion obtenida correctamente",
            "content": observacion
        }, status=status.HTTP_200_OK)

    def post(self, request: Request):
        anio = request.data.get('anio')
        mes = request.data.get('mes')
        observacion = request.data.get('observacion')
        usuario = request.user.username

        if not anio or not mes or not observacion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        observacion = S42InsertarCapacitacionObservacion(anio, mes, observacion, usuario)

        return Response(data={
            "message": "Observación insertada correctamente",
            "content": observacion
        }, status=status.HTTP_201_CREATED)


    def put(self, request: Request, id_observacion: int):
        observacion = request.data.get('observacion')
        usuario = request.user.username

        if observacion == None:
            return Response(data={
                "message": "Debe de ingresar observación",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        S42UpdateCapacitacionObservacion(id_observacion, observacion, usuario)

        return Response(data={
            "message": "Observación actualizada correctamente",
            "content": None
        }, status=status.HTTP_200_OK)

def create_excel_capacitacion(capacitaciones):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Capacitación"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    sheet.append(["Fecha", "Tema", "Modalidad", "Capacitador", "Empresas", "Lugar", "Cantidad", "Observación"])
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for capacitacion in capacitaciones:
        fecha_original = capacitacion.get("D_Capacita_Fecha")
        fecha_formateada = fecha_original
        
        if fecha_original:
            try:
                if isinstance(fecha_original, str):
                    fecha_obj = datetime.strptime(fecha_original, "%Y-%m-%d")
                    fecha_formateada = fecha_obj.strftime("%d/%m/%Y")
                elif hasattr(fecha_original, 'strftime'):
                    fecha_formateada = fecha_original.strftime("%d/%m/%Y")
            except:
                fecha_formateada = fecha_original
        
        sheet.append([fecha_formateada, capacitacion.get("N_Capacita_Tema"), capacitacion.get("N_Capacita_Modalidad"), capacitacion.get("N_Capacita_Capacitador"), capacitacion.get("N_Capacita_Empresas"), capacitacion.get("N_Capacita_Lugar"), capacitacion.get("Q_Capacita_Cantidad"), capacitacion.get("T_Capacita_Observ")])

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
   

   
@api_view(['POST'])
@permission_classes([AllowAny])
def DownloadCapacitacionController(request):
    try:
        anio = request.data.get('anio')
        mes = request.data.get('mes')

        if not anio:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        if not mes:
            capacitaciones= S42SelectCapacitacion("04", anio)
        else:
            capacitaciones = S42SelectCapacitacion("02", anio, mes)

        output = create_excel_capacitacion(capacitaciones)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="capacitacion.xlsx"'
        return response
    except Exception as e:
        return Response(data={
            "message": str(e),
            "content": None
        }, status=status.HTTP_400_BAD_REQUEST)


class SelectSenializaController(APIView):
    permission_classes = (IsAuthenticated,)

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get(self, request: Request):
        opcion = request.query_params.get('opcion')
        valor1 = request.query_params.get('valor1')
        valor2 = request.query_params.get('valor2')

        if not opcion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        senializacion = SelectSenializa(opcion, valor1, valor2)

        return Response(data={
            "message": "Senializacion obtenida correctamente",
            "content": senializacion
        }, status=status.HTTP_200_OK)

    def post(self, request: Request, anio: int, mes: int):

        usuario = request.user.username
        senializaciones = request.data.get('senializaciones')

        if not anio or not mes or not senializaciones:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        # Construir XML
        root = ET.Element('Senializaciones')
        for item in senializaciones:
            sen = ET.SubElement(root, 'Senializacion')            
            ET.SubElement(sen, 'C_Senializa_Indicador').text = str(item['C_Senializa_Indicador'])
            ET.SubElement(sen, 'Q_Senializa_Cantidad').text = str(item['Q_Senializa_Cantidad'])            
        
        xml_senializaciones = ET.tostring(root, encoding='unicode')

        S42InsertarSenializaciones(anio, mes, usuario, xml_senializaciones)

        return Response(data={
            "message": "Senializaciones insertadas correctamente",
            "content": None
        }, status=status.HTTP_200_OK)


class SelectSenializaIndicadorController(APIView):
    permission_classes = (IsAuthenticated,)

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get(self, request: Request):
        opcion = request.query_params.get('opcion')
        valor1 = request.query_params.get('valor1')
        valor2 = request.query_params.get('valor2')

        if not opcion:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        senializacion_indicador = SelectSenializaIndicador(opcion, valor1, valor2)

        return Response(data={
            "message": "Senializacion indicador obtenida correctamente",
            "content": senializacion_indicador
        }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def DownloadSenializacionesController(request):
    try:
        anio = request.data.get('anio')
        mes = request.data.get('mes')

        if not anio:
            return Response(data={
                "message": "Faltan parámetros",
                "content": None
            }, status=status.HTTP_400_BAD_REQUEST)

        if not mes:
            senializaciones= SelectSenializa("03", anio)
        else:
            senializaciones = SelectSenializa("02", anio, mes)

        output = create_excel_senializaciones(senializaciones)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="senializaciones.xlsx"'
        return response
    except Exception as e:
        return Response(data={
            "message": str(e),
            "content": None
        }, status=status.HTTP_400_BAD_REQUEST)


def create_excel_senializaciones(senializaciones):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Senializaciones"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    sheet.append(["Año", "Mes", "Indicador", "Cantidad", "Unidad de medida"])
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for senializacion in senializaciones:
        anio = senializacion.get("M_Senializa_Anio")
        mes = getMonthName(senializacion.get("M_Senializa_Mes"))
        indicador = senializacion.get("N_Senializa_Indicador")
        cantidad = senializacion.get("Q_Senializa_Cantidad")
        unidad_medida = senializacion.get("N_unimed_desc")
        
        sheet.append([anio, mes, indicador, cantidad, unidad_medida])

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output