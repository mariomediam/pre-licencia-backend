import os
import io
import requests
import urllib3
import urllib

import numpy as np
from statistics import mean, stdev

import pandas as pd
from sqlalchemy import create_engine


from django.shortcuts import render
from django.template.loader import get_template
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Sum, F, Max
from decimal import Decimal
from django.db import transaction


from rest_framework import status
from rest_framework.generics import RetrieveAPIView, CreateAPIView, UpdateAPIView, RetrieveUpdateAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.views import APIView

from datetime import datetime
from decimal import Decimal

import pdfkit
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl


from .siaf import *
from app_deploy.views import number_to_word_currency, BuscarSunat
from .models import Sincronizacion, RegistroSincronizacion, ProyectoInversion, ProgramacionProyectoInversion
from .serializers import ProyectoInversionSerializer


RESOURCE_ID = "35bdc5b5-017c-42c1-ba20-8820bf1248b7"
PATH_TEMP = os.path.join(os.path.dirname(__file__),"temp")

# c_proinv_codigo	=> CUI
# n_proinv_nombre	=> Proyecto
# q_prgpro_financ	=> Programacion inicial (a)
# q_prgpro_caida	=> Caídas (b)
# q_prgpro_increm	=> Incrementos (c) 
# q_prgpro_financ - q_prgpro_caida + q_prgpro_increm	=> Proyección ajustada (d) = (a) - (b) + ©
# MONTO_DEVENGADO_MES	=> Ejecución (e) 
# (q_prgpro_financ - q_prgpro_caida + q_prgpro_increm) / MONTO_DEVENGADO_MES	=> Avance financiero
# p_prgpro_fisica	=> Avance físico
# q_prgpro_riesgo	=> Riesgo a fin de mes
# t_prgpro_coment	=> Comentario

RENAME_COLUMNS = {
    "GASTO_MENSUAL": {
        "c_proinv_codigo": "CUI",
        "n_proinv_nombre": "Proyecto",
        "q_prgpro_financ": "Programacion inicial (a)",
        "q_prgpro_caida": "Caídas (b)",
        "q_prgpro_increm": "Incrementos (c)",
        "q_prgpro_financ - q_prgpro_caida + q_prgpro_increm": "Proyección ajustada (d) = (a) - (b) + ©",
        "MONTO_DEVENGADO_MES": "Ejecución (e)",
        "(q_prgpro_financ - q_prgpro_caida + q_prgpro_increm) / MONTO_DEVENGADO_MES": "Avance financiero",
        "p_prgpro_fisica": "Avance físico",
        "q_prgpro_riesgo": "Riesgo a fin de mes",
        "t_prgpro_coment": "Comentario"
    }
}


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Create your views here.
class MaestroDocumentoView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):        
        data = sf_listar_maestro_documento()
        return Response(
            data={"message": "Lista de documentos", "content": data},
            status=status.HTTP_200_OK,
        )
    
class PersonaView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request: Request):        
        filtro = request.data.get("filtro")
        if not filtro:
            return Response(
                data={"message": "Filtro no encontrado"},
                status=status.HTTP_400_BAD_REQUEST,
            )    

        data = sf_seleccionar_persona(filtro)        
        return Response(
            data={"message": "Lista de personas", "content": data},
            status=status.HTTP_200_OK,
        )
    

class ProveedorSIGAView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request: Request):        
        filtro = request.data.get("filtro")
        if not filtro:
            return Response(
                data={"message": "Filtro no encontrado"},
                status=status.HTTP_400_BAD_REQUEST,
            )    

        data = sf_seleccionar_proveedor(filtro)        
        return Response(
            data={"message": "Lista de proveedores SIGA", "content": data},
            status=status.HTTP_200_OK,
        )
    

class SeleccionarExpedienteFase(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):        
        ano_eje = request.query_params.get("anio")
        expediente = request.query_params.get("expediente")
        ciclo = request.query_params.get("ciclo")
        fase = request.query_params.get("fase")

        filters = {
            "ano_eje": ano_eje,
            "expediente": expediente,
            "ciclo": ciclo,
            "fase": fase,
        }

        if not ano_eje or not expediente or not ciclo or not fase:
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )    

        data = sf_seleccionar_expediente_fase(**filters)        
        return Response(
            data={"message": "Expediente fase", "content": data},
            status=status.HTTP_200_OK,
        )
    
# sf_seleccionar_expediente_secuencia    

class SeleccionarExpedienteSecuencia(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):        
        ano_eje = request.query_params.get("anio")
        expediente = request.query_params.get("expediente")
        secuencia = request.query_params.get("secuencia")
        correlativo = request.query_params.get("correlativo")

        filters = {
            "ano_eje": ano_eje,
            "expediente": expediente,
            "secuencia": secuencia,
            "correlativo": correlativo,
        }

        if not ano_eje or not expediente or not secuencia or not correlativo:
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )    

        data = sf_seleccionar_expediente_secuencia(**filters)        
        return Response(
            data={"message": "Expediente secuencia", "content": data},
            status=status.HTTP_200_OK,
        )
    
def get_context_secuencia(data, retentions):
    total_retentions = 0
    for retention in retentions:
        total_retentions += retention["value"]

    total_retentions_decimal = Decimal(str(total_retentions))
    
    monto_final = data["MONTO_NACIONAL"] - total_retentions_decimal
    
    data["retentions"] = retentions
    data["total_retentions"] = total_retentions_decimal
    data["monto_final"] = monto_final
    data["monto_final_text"] = number_to_word_currency(monto_final).upper()
    return data

def correct_supplier_name(data):    
    nro_ruc = data.get("RUC", "")
    if nro_ruc:
        filters = {
            "nro_ruc": nro_ruc
        }
        
        result = sf_seleccionar_sigamef_contratista(**filters)
        if result:
            data["NOMBRE"] = result.get("NOMBRE_PROV", "")
    return data

    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def DownloadFormatoDevengadoController(request):    
    try:

        ano_eje = request.data.get("anio")
        expediente = request.data.get("expediente")
        secuencia = request.data.get("secuencia")
        correlativo = request.data.get("correlativo")
        retentions = request.data.get("retentions")
        login = request.user   
        fecha_print = datetime.now().strftime('%d/%m/%Y %I:%M:%S %p')                  

        if ano_eje and expediente and secuencia and correlativo and retentions:

            # ********************* INCIO HEADER PDF ********************* #                            
            
            template = get_template('formato-devengado-header.html')
            context = {"ANO_EJE" : ano_eje, "EXPEDIENTE" : expediente,}
            
            options = {
                'encoding': "UTF-8",                           
            }  

            page_header = template.render(context = context)
            header_template_path = os.path.join(settings.MEDIA_ROOT, 'reque-header' + str(uuid.uuid4()) + '.html')
            
            text_file = open(header_template_path, "w", encoding='utf-8')            
            text_file.write(page_header)
            text_file.close()

            # ********************* INCIO FOOTER PDF ********************* #
            
            template = get_template('formato-devengado-footer.html')

            context = {"login": login, "fecha_print": fecha_print}

            page_footer = template.render(context = context)
            
            footer_template_path = os.path.join(settings.MEDIA_ROOT, 'reque-footer' + str(uuid.uuid4()) + '.html')
            
            text_file = open(footer_template_path, "w", encoding='utf-8')            
            text_file.write(page_footer)
            text_file.close()

            

            # ********************* INCIO MAIN PDF ********************* #  
            filters = {
                "ano_eje": ano_eje,
                "expediente": expediente,
                "secuencia": secuencia,
                "correlativo": correlativo,
            }          

            data = sf_seleccionar_expediente_secuencia(**filters)  

            if not data["NOMBRE"]:
                data = correct_supplier_name(data)
            
            context = get_context_secuencia(data, retentions)   

            template = get_template('formato-devengado.html')            
            html = template.render(context = context)    
            
            options = {
                'page-size': 'A4',
                'margin-top': '2.0in',
                'margin-right': '0.25in',
                'margin-bottom': '0.50in',
                'margin-left': '0.25in',
                'encoding': "UTF-8",                           
                'footer-html': footer_template_path,    
                'header-html': header_template_path,
                'footer-font-size':'7',
                'footer-right': 'Pagina [page] de [topage]',
                           
            }        
                    
            file_generate = pdfkit.from_string(html, False, options=options)
            response = HttpResponse(file_generate, content_type="application/pdf")            
            file_name_download = "requerimiento{}-{}-{}-{}.pdf".format(ano_eje, expediente, secuencia, correlativo)            
            response['Content-Disposition'] = "attachment; filename={}".format(file_name_download)

            # ********************* FIN GENERANDO PDF ********************* #
            os.remove(header_template_path)
            os.remove(footer_template_path)


            return response

        else:
            return Response(
                data={
                    "message": "Faltan parametros",
                    "content": None,
                },
                status=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        
        return Response(
            data={"message": str(e), "content": None},
            status=status.HTTP_404_NOT_FOUND,
        )         
    

class ProcesoActualizarRegistroView(CreateAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request: Request):        
        ano_eje = request.data.get("anio")
        expediente = request.data.get("expediente")

        print("Ingreso *************")

        if not ano_eje or not expediente:
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )    
        
        try:
            filters = {
            "ano_eje": ano_eje,
            "expediente": expediente,
            }

            sf_proceso_actualizar_01_registro(**filters)        
            return Response(
                data={"message": "Proceso actualizar registro", "content": "Expediente migrado exitosamente"},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                data={"message": str(e), "content": None},
                status=status.HTTP_404_NOT_FOUND,
            )
        
class BuscarCartaOrdenView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request: Request):        
        cod_doc = request.data.get("codigo")
        num_doc = request.data.get("numero")

        if not cod_doc or not num_doc:
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )    

        try:
            filters = {
                "cod_doc": cod_doc,
                "num_doc": num_doc,
            }

            data = sf_buscar_carta_orden(**filters)        
            return Response(
                data={"message": "Carta orden", "content": data},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                data={"message": str(e), "content": None},
                status=status.HTTP_404_NOT_FOUND,
            )
        
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def DownloadCartaOrdenFideicomisoController(request):    
    try:
        cartas = request.data.get("cartas")

        if not cartas:
            return Response(
                data={"message": "Faltan parametros"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        login = request.user

        cartas = get_domicilio_carta_orden(cartas)

        output = create_excel_carta_orden(cartas)
        
        # Crear la respuesta HTTP con el contenido del archivo Excel
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="reporte.xlsx"'
        return response
    
    except Exception as e:
        return Response(
            data={"message": str(e), "content": None},
            status=status.HTTP_404_NOT_FOUND,
        )

def create_excel_carta_orden(cartas):
    # Crear un nuevo libro de Excel y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

   # -- Fila 1: Título principal --
    ws.merge_cells("A1:H1")
    ws["A1"] = "FIDEICOMISO MUNICIPALIDAD PROVINCIAL DE PIURA - MINAM"    

    # -- Fila 2: Subtítulo (RUC) --
    ws.merge_cells("A2:H2")
    ws["A2"] = "RUC 20154477374"

    # quiero que la celda a1 tenga color de fondo gris y color de texto blanco
    for cell in ["A1", "A2"]:
        ws[cell].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        ws[cell].font = Font(color="FFFFFF")

    # -- Fila 3: Encabezados de columnas --
    headers = [
        "OFICIO",            # A
        "N° DE ORDEN PAGO",  # B
        "BENEFICIARIO / MONTO EN LETRAS",  # C
        "N° DE CUENTA/CCI",  # D
        "TIPO DE CUENTA",    # E
        "BANCO",             # F
        "MONTO",             # G
        "CONCEPTO DE PAGO"   # H
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Ajustar anchos de columna
    col_widths = [8, 18, 40, 25, 15, 15, 12, 50]  # A-H
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Fila a partir de la cual empezaremos a escribir los datos
    start_row = 4

    for carta in cartas:
        # Por cada 'carta', vamos a usar 5 filas
        # -> Fusionar las celdas de OFICIO (A) para 4 filas
        ws.merge_cells(start_row=start_row, end_row=start_row+3, 
                       start_column=1, end_column=1)
        # -> Fusionar las celdas de ORDEN DE PAGO (B) para 4 filas
        ws.merge_cells(start_row=start_row, end_row=start_row+3, 
                       start_column=2, end_column=2)
        # -> Fusionar las celdas de CONCEPTO DE PAGO (H) para 4 filas
        ws.merge_cells(start_row=start_row, end_row=start_row+3, 
                       start_column=8, end_column=8)

        # OFICIO (ejemplo fijo o en blanco)
        oficio_value = ""  # o "49"
        ws.cell(row=start_row, column=1, value=oficio_value)
        ws.cell(row=start_row, column=1).alignment = Alignment(vertical="top")

        # ORDEN DE PAGO (ejemplo fijo o en blanco)
        orden_pago_value = carta.get("NUM_DOC", "")
        ws.cell(row=start_row, column=2, value=orden_pago_value)
        ws.cell(row=start_row, column=2).alignment = Alignment(vertical="top")

        # CONCEPTO DE PAGO
        glosa = carta.get("GLOSA", "")
        ws.cell(row=start_row, column=8, value=glosa)
        ws.cell(row=start_row, column=8).alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

        # Extraemos datos del "carta"
        proveedor = carta.get("NOMBRE_PROVEEDOR", "")
        ruc = carta.get("RUC", "")
        monto_numerico = carta.get("MONTO_NACIONAL", 0.0)
        monto_en_letras = f"SON: {number_to_word_currency(monto_numerico).upper()}"
        domicilio = f"DOMICIO LEGAL: {carta.get('DOMICILIO', '')}" 

        # N° CUENTA / CCI
        cci = carta.get("CCI", "")
        # TIPO DE CUENTA (ejemplo: "CORRIENTE")
        tipo_cuenta = carta.get("NOMBRE_CUENTA_CTE", "")
        # BANCO
        banco = carta.get("NOMBRE_BANCO", "")
        # MONTO (col G)
        monto = monto_numerico

        # -- Llenar 4 filas (columnas C-G) --

        # 1) Fila principal: Proveedor (col C), CCI (D), tipo cuenta (E), banco (F), monto (G)
        row_proveedor = start_row
        ws.cell(row=row_proveedor, column=3, value=proveedor)         
        ws.cell(row=row_proveedor, column=4, value=cci)
        ws.cell(row=row_proveedor, column=5, value=tipo_cuenta)
        ws.cell(row=row_proveedor, column=6, value=banco)
        celda_monto = ws.cell(row=row_proveedor, column=7, value=monto)
        celda_monto.number_format = '#,##0.00'


        # 2) Fila: Monto en letras (col C)
        row_monto_letras = start_row + 1        
        ws.merge_cells(F"C{row_monto_letras}:G{row_monto_letras}")
        ws.cell(row=row_monto_letras, column=3, value=monto_en_letras)

        # 3) Fila: DNI/RUC (col C)
        row_ruc = start_row + 2
        ws.merge_cells(F"C{row_ruc}:G{row_ruc}")
        ws.cell(row=row_ruc, column=3, value=f"DNI/RUC: {ruc}")

        # 4) Fila: Domicilio (col C)
        row_domicilio = start_row + 3
        ws.merge_cells(F"C{row_domicilio}:G{row_domicilio}")
        ws.cell(row=row_domicilio, column=3, value=domicilio)

        # Ajustes de alineación / wrap text en las celdas de col C
        for r in range(start_row, start_row + 4):
            cell_c = ws.cell(row=r, column=3)
            cell_c.alignment = Alignment(wrap_text=True, vertical="top")

        for col in range(4, 8):
            ws.cell(row=start_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Ajustar altura de las filas
        for r in range(start_row, start_row + 4):
            # Cada fila un poco más alta para soportar wrap
            ws.row_dimensions[r].height = 22

        # Avanzar 4 filas para la siguiente 'carta'
        start_row += 4
    
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=3)
    ws.cell(row=start_row, column=1).alignment = Alignment(wrap_text=True, vertical="bottom")

    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=4, end_column=6)
    ws.cell(row=start_row, column=4).alignment = Alignment(wrap_text=True, vertical="bottom")

    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=7, end_column=8)
    ws.cell(row=start_row, column=7).alignment = Alignment(wrap_text=True, vertical="center")

    cell_values = {
    1: "2/ Firma autorizada",
    4: "2/ Firma autorizada",
    7: ("1/ Toda la información sobre los datos de los fideicomisarios indicada en las instrucciones de transferencia, es de responsabilidad única y exclusiva del Fideicomitente. Estas transferencias de fondos, están relacionadas a la finalidad del fideicomiso.\n"
        "2/ Dobles firmas registradas en el Banco de la Nación.")
    }

    for col, value in cell_values.items():
        celda = ws.cell(row=start_row, column=col, value=value)
        celda.font = Font(size=8)        

    ws.row_dimensions[start_row].height = 120

    # Aplicar borde a toda la tabla (encabezados, datos y fila de sumatoria)
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Agregar una fila de sumatoria al final (excepto Avance financiero y Avance físico)
    suma_dict = {}
    for col_name in ws[1]:
        header = col_name.value
        if header not in ['Avance financiero', 'Avance físico', 'Proyecto', 'Comentario', 'CUI']:
            idx = col_name.column
            col_letter = col_name.column_letter
            # Sumar solo si la columna es numérica
            try:
                suma = sum([cell.value for cell in ws[col_letter][1:] if isinstance(cell.value, (int, float))])
                suma_dict[header] = suma
            except:
                suma_dict[header] = ''
        else:
            suma_dict[header] = ''
    # Crear la fila de sumatoria
    sum_row = []
    for col_name in ws[1]:
        header = col_name.value
        if header == 'Proyecto':
            sum_row.append('TOTAL')
        elif header in suma_dict:
            val = suma_dict[header]
            # Si el valor es numérico, asegúrate de que sea float
            if isinstance(val, (int, float)):
                sum_row.append(float(val))
            else:
                sum_row.append('')
        else:
            sum_row.append('')
    ws.append(sum_row)

    # Aplicar borde a la fila de sumatoria
    thin_side = Side(style='thin', color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.border = thin_border

    # Aplicar formato numérico y porcentaje a la fila de sumatoria (igual que a las filas de datos)
    last_row = ws.max_row
    for col_name in ws[1]:
        header = col_name.value
        idx = col_name.column
        cell = ws.cell(row=last_row, column=idx)
        if header in columnas_moneda:
            cell.number_format = '#,##0.00'
        elif header in columnas_porcentaje:
            cell.number_format = '0.00%'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # Reiniciar la posición al inicio
    return output

def get_domicilio_carta_orden(cartas):

    domicilios_cache = {}
    
    for carta in cartas:
        ruc = carta.get("RUC", "")
        if not ruc:
            carta["DOMICILIO"] = ""
            continue
        
        if ruc not in domicilios_cache:
            resultado = BuscarSunat(ruc)            
            domicilio = f"{resultado.get('desc_tipvia', '').strip()} {resultado.get('ddp_nomvia', '').strip()} {resultado.get('ddp_numer1', '')} {resultado.get('ddp_nomzon', '').strip()} {resultado.get('desc_dep', '').strip()} - {resultado.get('desc_prov', '').strip()} - {resultado.get('desc_dist', '').strip()}".strip()
            domicilios_cache[ruc] = domicilio
        else:
            domicilio = domicilios_cache[ruc]
        
        # Asignar el domicilio obtenido a la carta
        carta["DOMICILIO"] = domicilio
    
    return cartas


class SincroGastoDiario(RetrieveAPIView):
   permission_classes = [IsAuthenticated]

   def get(self, request: Request):
      try:
        # Crear el objeto Sincronizacion antes de llamar a la API
        sinc = Sincronizacion.objects.create()

        ultima_actualizacion = getLastDataUpdate()
        
        # Obtener datos de la API
        data = getGastoDiario()

        # Pasar el objeto sincronizacion a la función saveGastoDiario
        saveGastoDiario(data, sinc, ultima_actualizacion)

        return Response(
            data={"message": "Sincronización completada exitosamente", "content": sinc.idSincro},
            status=status.HTTP_200_OK,
        )
                  
      except Exception as e:               
        return Response(
            data={"message": str(e), "content": None},
            status=status.HTTP_404_NOT_FOUND,
        )     
        

def getGastoDiario():   

      sql = (
         'SELECT * '
         'FROM "35bdc5b5-017c-42c1-ba20-8820bf1248b7" '
         "WHERE \"SEC_EJEC\" = '301529' "            
      )

      resp = requests.get(
         'https://api.datosabiertos.mef.gob.pe/DatosAbiertos/v1/datastore_search_sql',
         params={'sql': sql}
      )
      
      return resp.json()


def saveGastoDiario(gasto_diario, sinc, ultima_actualizacion):
    """
    Guarda en BD los registros de gasto diario traídos desde la API.
    Usa el objeto Sincronizacion creado previamente.
    """
    try:
        registros = gasto_diario.get('records', [])
        fila_error = None

        # Insertar cada fila en RegistroSincronizacion
        for fila in registros:
            fila_error = fila
            RegistroSincronizacion.objects.create(
                sincronizacion=sinc,
                ano_eje=fila.get('ANO_EJE'),
                mes_eje=fila.get('MES_EJE'),
                nivel_gobierno=fila.get('NIVEL_GOBIERNO'),
                nivel_gobierno_nombre=fila.get('NIVEL_GOBIERNO_NOMBRE'),
                sector=fila.get('SECTOR'),
                sector_nombre=fila.get('SECTOR_NOMBRE'),
                pliego=fila.get('PLIEGO'),
                pliego_nombre=fila.get('PLIEGO_NOMBRE'),
                sec_ejec=fila.get('SEC_EJEC'),
                ejecutora=fila.get('EJECUTORA'),
                ejecutora_nombre=fila.get('EJECUTORA_NOMBRE'),
                departamento_ejecutora=fila.get('DEPARTAMENTO_EJECUTORA'),
                departamento_ejecutora_nombre=fila.get('DEPARTAMENTO_EJECUTORA_NOMBRE'),
                provincia_ejecutora=fila.get('PROVINCIA_EJECUTORA'),
                provincia_ejecutora_nombre=fila.get('PROVINCIA_EJECUTORA_NOMBRE'),
                distrito_ejecutora=fila.get('DISTRITO_EJECUTORA'),
                distrito_ejecutora_nombre=fila.get('DISTRITO_EJECUTORA_NOMBRE'),
                sec_func=fila.get('SEC_FUNC'),
                programa_ppto=fila.get('PROGRAMA_PPTO'),
                programa_ppto_nombre=fila.get('PROGRAMA_PPTO_NOMBRE'),
                tipo_act_proy=fila.get('TIPO_ACT_PROY'),
                tipo_act_proy_nombre=fila.get('TIPO_ACT_PROY_NOMBRE'),
                producto_proyecto=fila.get('PRODUCTO_PROYECTO'),
                producto_proyecto_nombre=fila.get('PRODUCTO_PROYECTO_NOMBRE'),
                actividad_accion_obra=fila.get('ACTIVIDAD_ACCION_OBRA'),
                actividad_accion_obra_nombre=fila.get('ACTIVIDAD_ACCION_OBRA_NOMBRE'),
                funcion=fila.get('FUNCION'),
                funcion_nombre=fila.get('FUNCION_NOMBRE'),
                division_funcional=fila.get('DIVISION_FUNCIONAL'),
                division_funcional_nombre=fila.get('DIVISION_FUNCIONAL_NOMBRE'),
                grupo_funcional=fila.get('GRUPO_FUNCIONAL'),
                grupo_funcional_nombre=fila.get('GRUPO_FUNCIONAL_NOMBRE'),
                meta=fila.get('META'),
                finalidad=fila.get('FINALIDAD'),
                meta_nombre=fila.get('META_NOMBRE'),
                departamento_meta=fila.get('DEPARTAMENTO_META'),
                departamento_meta_nombre=fila.get('DEPARTAMENTO_META_NOMBRE'),
                fuente_financiamiento=fila.get('FUENTE_FINANCIAMIENTO'),
                fuente_financiamiento_nombre=fila.get('FUENTE_FINANCIAMIENTO_NOMBRE'),
                rubro=fila.get('RUBRO'),
                rubro_nombre=fila.get('RUBRO_NOMBRE'),
                tipo_recurso=fila.get('TIPO_RECURSO'),
                tipo_recurso_nombre=fila.get('TIPO_RECURSO_NOMBRE'),
                categoria_gasto=fila.get('CATEGORIA_GASTO'),
                categoria_gasto_nombre=fila.get('CATEGORIA_GASTO_NOMBRE'),
                tipo_transaccion=fila.get('TIPO_TRANSACCION'),
                generica=fila.get('GENERICA'),
                generica_nombre=fila.get('GENERICA_NOMBRE'),
                subgenerica=fila.get('SUBGENERICA'),
                subgenerica_nombre=fila.get('SUBGENERICA_NOMBRE'),
                subgenerica_det=fila.get('SUBGENERICA_DET'),
                subgenerica_det_nombre=fila.get('SUBGENERICA_DET_NOMBRE'),
                especifica=fila.get('ESPECIFICA'),
                especifica_nombre=fila.get('ESPECIFICA_NOMBRE'),
                especifica_det=fila.get('ESPECIFICA_DET'),
                especifica_det_nombre=fila.get('ESPECIFICA_DET_NOMBRE'),
                monto_pia=fila.get('MONTO_PIA'),
                monto_pim=fila.get('MONTO_PIM'),
                monto_certificado=fila.get('MONTO_CERTIFICADO'),
                monto_comprometido_anual=fila.get('MONTO_COMPROMETIDO_ANUAL'),
                monto_comprometido=fila.get('MONTO_COMPROMETIDO'),
                monto_devengado=fila.get('MONTO_DEVENGADO'),
                monto_girado=fila.get('MONTO_GIRADO'),
            )
        
        # Éxito
        sinc.comentarios = f'Sincronización exitosa ({len(registros)} registros).'
        sinc.exitoso = True
    except Exception as e:
        print("************** error ***********")
        print(fila_error)
        # Cualquier excepción se graba en comentarios
        sinc.comentarios = f'Error durante sincronización: {e}'
        sinc.exitoso = False        
    finally:
        # Marcar fecha_fin y guardar
        #sinc.fecha_fin = timezone.now()
        sinc.fecha_fin = datetime.now()
        sinc.ultima_actualizacion = ultima_actualizacion
        sinc.save()


def getLastDataUpdate(resource_id = RESOURCE_ID):
   
   url = "https://datosabiertos.mef.gob.pe/Rest/PortalWebRecursoDetalle/v1.0/getRecursoDetalle"
   
   # Datos para el body de la petición
   payload = {
       "dataset": "presupuesto-y-ejecucion-de-gasto",
       "id_recurso": resource_id
   }
   
   # Headers para la petición
   headers = {
       'Content-Type': 'application/json'
   }

   try:
       response = requests.post(
           url,
           json=payload,
           headers=headers,
           verify=False  # Deshabilitar verificación SSL
       )
       response.raise_for_status()  # Verificar si hay errores HTTP
       
       # Obtener el JSON de la respuesta
       data = response.json()
       
       # Obtener el valor de resource_last_modified
       last_modified = data.get('resource_detail', {}).get('resource_last_modified')
              
       return last_modified
       
   except requests.exceptions.SSLError as e:
       print(f"Error SSL: {e}")
       raise Exception(f"Error de SSL: {str(e)}")
   except requests.exceptions.RequestException as e:
       print(f"Error en la petición: {e}")
       raise Exception(f"Error en la petición: {str(e)}")
   except Exception as e:
       print(f"Error inesperado: {e}")
       raise Exception(f"Error inesperado: {str(e)}")


def getLastSincro():
    return Sincronizacion.objects.filter(
    exitoso=True
    ).order_by('-idSincro').first()

def getLastSincroByAnoEjecutora(ano_eje, sec_ejec):
    """
    Obtiene la última sincronización exitosa para un año y ejecutora específicos.
    
    Args:
        ano_eje (str): Año de ejecución
        sec_ejec (str): Código de la ejecutora
        
    Returns:
        Sincronizacion: Objeto de la última sincronización exitosa o None si no existe
    """
    return Sincronizacion.objects.filter(
        registros__ano_eje=ano_eje,
        registros__sec_ejec=sec_ejec,
        exitoso=True
    ).order_by('-idSincro').first()

class UltimaSincronizacionView(RetrieveAPIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        try:
            ano_eje = request.query_params.get('ano_eje')
            sec_ejec = request.query_params.get('sec_ejec')

            if not ano_eje or not sec_ejec:
                return Response(
                    data={"message": "Faltan parámetros ano_eje o sec_ejec"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            sincronizacion = Sincronizacion.obtener_ultima_sincronizacion_completa(ano_eje, sec_ejec)

            if not sincronizacion:
                return Response(
                    data={"message": "No se encontró sincronización exitosa", "content": None},
                    status=status.HTTP_200_OK
                )

            return Response(
                data={
                    "message": "Última sincronización encontrada",
                    "content": {
                        "idSincro": sincronizacion.idSincro,
                        "fecha_inicio": sincronizacion.fecha_inicio,
                        "fecha_fin": sincronizacion.fecha_fin,
                        "ultima_actualizacion": sincronizacion.ultima_actualizacion,
                        "comentarios": sincronizacion.comentarios,
                        "exitoso": sincronizacion.exitoso
                    }
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                data={"message": str(e), "content": None},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ProgProyectosInversionMensualView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        try:
            ano_eje = request.query_params.get('ano_eje')
            mes_eje = request.query_params.get('mes_eje')
            sec_ejec = request.query_params.get('sec_ejec')
            c_proinv_codigo = request.query_params.get('c_proinv_codigo')

            if not all([ano_eje, mes_eje, sec_ejec]):
                return Response(
                    data={"message": "Faltan parámetros requeridos", "content": None},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
          
            
            # Obtener la última sincronización
            ultima_sincro = Sincronizacion.obtener_ultima_sincronizacion_completa(ano_eje, sec_ejec)
            
            if not ultima_sincro:
                return Response(
                    data={"message": "No se encontró sincronización exitosa", "content": []},
                    status=status.HTTP_200_OK
                )
            
            filters = {
                "ano_eje": ano_eje,
                "mes_eje": mes_eje,
                "sec_ejec": sec_ejec,
                "sincronizacion_id": ultima_sincro.idSincro
            }
            
            result = select_protectos_con_gasto_mensual(**filters)

            if c_proinv_codigo:
                result = [item for item in result if item.get('c_proinv_codigo').strip()  == c_proinv_codigo.strip()]

            return Response(
                data={
                    "message": "Proyectos de inversión encontrados",
                    "content": result
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                data={"message": str(e), "content": None},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        


def obtener_producto_proyecto(ano_eje: float, producto_proyecto: float) -> dict:
    """
    Obtiene la información de un producto/proyecto específico para un año dado.
    
    Args:
        ano_eje (float): Año de ejecución
        producto_proyecto (float): Código del producto/proyecto
        
    Returns:
        dict: Diccionario con la información del producto/proyecto o None si no se encuentra
    """
    registro = RegistroSincronizacion.objects.filter(
        ano_eje=ano_eje,
        producto_proyecto=producto_proyecto
    ).values('ano_eje', 'producto_proyecto', 'producto_proyecto_nombre').first()
    
    return registro


# CREAR UNA VISTA GET PARA OBTENER EL PRODUCTO PROYECTO
class ProductoProyectoNombreView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        try:
            ano_eje = request.query_params.get('ano_eje')
            producto_proyecto = request.query_params.get('producto_proyecto')

            if not ano_eje or not producto_proyecto:
                return Response(
                    data={"message": "Faltan parámetros requeridos", "content": None},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            result = obtener_producto_proyecto(ano_eje, producto_proyecto)
            return Response(
                data={
                    "message": "Proyectos de inversión encontrados",
                    "content": result
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                data={"message": str(e), "content": None},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def obtener_ultima_sincronizacion_por_anio(ano_eje: float) -> int:
    """
    Obtiene el ID de la última sincronización exitosa para un año específico.
    
    Args:
        ano_eje (float): Año de ejecución
        
    Returns:
        int: ID de la última sincronización exitosa o None si no se encuentra
    """
    ultima_sincronizacion = Sincronizacion.objects.filter(
        registros__ano_eje=ano_eje,
        exitoso=True
    ).order_by('-idSincro').first()
    
    return ultima_sincronizacion.idSincro if ultima_sincronizacion else None


class UltimaSincronizacionAnioView(RetrieveAPIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        ano_eje = request.query_params.get('ano_eje')

        if not ano_eje:
            return Response(
                data={"message": "Faltan parámetros requeridos", "content": None},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        id_sincro = obtener_ultima_sincronizacion_por_anio(ano_eje)

        return Response(
            data={"message": "Última sincronización encontrada", "content": id_sincro},
            status=status.HTTP_200_OK
        )




def obtener_resumen_producto_proyecto(producto_proyecto: Decimal, id_sincro: int) -> dict:
    """
    Obtiene el resumen de montos para un producto/proyecto específico de una sincronización.
    
    Args:
        producto_proyecto (Decimal): Código del producto/proyecto
        id_sincro (int): ID de la sincronización
        
    Returns:
        dict: Diccionario con los montos resumidos o None si no se encuentra
    """
    resultado = RegistroSincronizacion.objects.filter(
        sincronizacion_id=id_sincro,
        producto_proyecto=producto_proyecto
    ).values(
        'ano_eje',
        'producto_proyecto'
    ).annotate(
        MONTO_PIA=Sum('monto_pia'),
        MONTO_PIM=Sum('monto_pim'),
        MONTO_CERTIFICADO=Sum('monto_certificado'),
        MONTO_COMPROMETIDO_ANUAL=Sum('monto_comprometido_anual'),
        MONTO_COMPROMETIDO=Sum('monto_comprometido'),
        MONTO_DEVENGADO=Sum('monto_devengado'),
        MONTO_GIRADO=Sum('monto_girado')
    ).order_by('ano_eje', 'producto_proyecto').first()
    
    return resultado



class ResumenProductoProyectoView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        producto_proyecto = request.query_params.get('producto_proyecto')
        ano_eje = request.query_params.get('ano_eje')

        if not producto_proyecto or not ano_eje:
            return Response(
                data={"message": "Faltan parámetros requeridos", "content": None},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        id_sincro = obtener_ultima_sincronizacion_por_anio(ano_eje)

        resumen = obtener_resumen_producto_proyecto(producto_proyecto, id_sincro)

        return Response(
            data={"message": "Resumen encontrado", "content": resumen},
            status=status.HTTP_200_OK
        )
    


# {
#     "ano_eje": 2025,
#     "c_proinv_codigo": "2331918",
#     "n_proinv_nombre": "MEJORAMIENTO DEL SERVICIO DE TRANSITABILIDAD PEATONAL Y VEHICULAR DE LA UPIS LOS ANGELES EN EL DISTRITO DE PIURA, PROVINCIA DE PIURA - PIURA",
#     "programacion": [
#         {
#             "m_prgpro_mes": 1,
#             "q_prgpro_financ": 100,
#             "p_prgpro_fisica": 1
#         },
#         {
#             "m_prgpro_mes": 2,
#             "q_prgpro_financ": 200,
#             "p_prgpro_fisica": 2
#         },
#         {
#             "m_prgpro_mes": 3,
#             "q_prgpro_financ": 300,
#             "p_prgpro_fisica": 3
#         },
#         {
#             "m_prgpro_mes": 4,
#             "q_prgpro_financ": 400,
#             "p_prgpro_fisica": 4
#         },
#         {
#             "m_prgpro_mes": 5,
#             "q_prgpro_financ": 500,
#             "p_prgpro_fisica": 5
#         },
#         {
#             "m_prgpro_mes": 6,
#             "q_prgpro_financ": 600,
#             "p_prgpro_fisica": 6
#         },
#         {
#             "m_prgpro_mes": 7,
#             "q_prgpro_financ": 700,
#             "p_prgpro_fisica": 7
#         },
#         {
#             "m_prgpro_mes": 8,
#             "q_prgpro_financ": 800,
#             "p_prgpro_fisica": 8
#         },
#         {
#             "m_prgpro_mes": 9,
#             "q_prgpro_financ": 900,
#             "p_prgpro_fisica": 9
#         },
#         {
#             "m_prgpro_mes": 10,
#             "q_prgpro_financ": 1000,
#             "p_prgpro_fisica": 10
#         },
#         {
#             "m_prgpro_mes": 11,
#             "q_prgpro_financ": 1100,
#             "p_prgpro_fisica": 11
#         },
#         {
#             "m_prgpro_mes": 12,
#             "q_prgpro_financ": 1200,
#             "p_prgpro_fisica": 34
#         }
#     ]
# }


# Dado el JSON anterior, se debe grabar el registro en las tablas ProyectoInversion y ProgramacionProyectoInversion

class ProyectoInversionView(CreateAPIView):    
    permission_classes = [IsAuthenticated]
    serializer_class = ProyectoInversionSerializer

    def post(self, request: Request):

        data = request.data
        data['c_usuari_login'] = request.user.username
        data['n_proinv_pc'] = request.META.get('REMOTE_ADDR')

        with transaction.atomic():
            serializer = self.serializer_class(data=data)

            if serializer.is_valid():
                serializer.save()
                proyecto = serializer.instance
                print(proyecto.c_proinv)

                for programacion in data['programacion']:
                    
                    programacion['proyecto'] = proyecto
                    programacion['c_usuari_login'] = request.user.username
                    programacion['n_prgpro_pc'] = request.META.get('REMOTE_ADDR')

                    ProgramacionProyectoInversion.objects.create(**programacion)

                return Response(serializer.data, status=status.HTTP_201_CREATED)
    
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, c_prgpro):
        try:
            # Obtener el usuario y la IP del cliente
            data = request.data.copy()
            data['c_usuari_login'] = request.user.username
            data['n_prgpro_pc'] = request.META.get('REMOTE_ADDR')
            
            resultado = actualizar_programacion_proyecto(c_prgpro, data)
            
            if resultado['success']:
                return Response(
                    data={
                        "message": resultado['message'],
                        "content": resultado['data']
                    },
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    data={
                        "message": resultado['message'],
                        "content": None
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            return Response(
                data={
                    "message": str(e),
                    "content": None
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

def obtener_programacion_proyecto(c_prgpro: int) -> dict:
    """
    Obtiene la información detallada de una programación de proyecto de inversión.
    
    Args:
        c_prgpro (int): ID de la programación del proyecto
        
    Returns:
        dict: Diccionario con la información de la programación y el proyecto asociado
    """
    resultado = ProgramacionProyectoInversion.objects.filter(
        c_prgpro=c_prgpro
    ).select_related(
        'proyecto',
        'estado'
    ).annotate(
        c_proinv=F('proyecto__c_proinv'),
        n_proinv_nombre=F('proyecto__n_proinv_nombre'),
        ano_eje=F('proyecto__ano_eje'),
        c_proinv_codigo=F('proyecto__c_proinv_codigo'),
        c_estado=F('estado__c_estado')

    ).values(
        'c_prgpro',
        'c_proinv',
        'm_prgpro_mes',
        'q_prgpro_financ',
        'p_prgpro_fisica',
        'q_prgpro_caida',
        'q_prgpro_increm',
        'q_prgpro_riesgo',
        't_prgpro_estsit',
        't_prgpro_coment',
        'c_estado',
        'c_usuari_login',
        'n_prgpro_pc',
        'd_prgpro_fecdig',
        'ano_eje',
        'c_proinv_codigo',
        'n_proinv_nombre'
    ).first()
    
    return resultado

class ProgramacionProyectoInversionView(RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = ProgramacionProyectoInversion.objects.all()

    def get(self, request, c_prgpro):
        try:
            resultado = obtener_programacion_proyecto(c_prgpro)
            
            if not resultado:
                return Response(
                    data={"message": "No se encontró la programación del proyecto"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            return Response(
                data={
                    "message": "Programación de proyecto encontrada",
                    "content": resultado
                },
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            return Response(
                data={"message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def put(self, request, c_prgpro):
        try:
            # Obtener el usuario y la IP del cliente
            data = request.data.copy()
            data['c_usuari_login'] = request.user.username
            data['n_prgpro_pc'] = request.META.get('REMOTE_ADDR')
            data['d_prgpro_fecdig'] = datetime.now()
            
            resultado = actualizar_programacion_proyecto(c_prgpro, data)
            
            if resultado['success']:
                return Response(
                    data={
                        "message": resultado['message'],
                        "content": resultado['data']
                    },
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    data={
                        "message": resultado['message'],
                        "content": None
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            return Response(
                data={
                    "message": str(e),
                    "content": None
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

def actualizar_programacion_proyecto(c_prgpro: int, data: dict) -> dict:
    """
    Actualiza un registro existente en la tabla PROGRAMACION_PROYECTOS_INVERSION.
    
    Args:
        c_prgpro (int): ID de la programación a actualizar
        data (dict): Diccionario con los datos a actualizar
        
    Returns:
        dict: Diccionario con el resultado de la operación
    """
    try:
        with transaction.atomic():
            # Buscar la programación existente
            programacion = ProgramacionProyectoInversion.objects.get(c_prgpro=c_prgpro)
            
            # Actualizar los campos
            programacion.m_prgpro_mes = data.get('m_prgpro_mes', programacion.m_prgpro_mes)
            programacion.q_prgpro_financ = data.get('q_prgpro_financ', programacion.q_prgpro_financ)
            programacion.p_prgpro_fisica = data.get('p_prgpro_fisica', programacion.p_prgpro_fisica)
            programacion.q_prgpro_caida = data.get('q_prgpro_caida', programacion.q_prgpro_caida)
            programacion.q_prgpro_increm = data.get('q_prgpro_increm', programacion.q_prgpro_increm)
            programacion.q_prgpro_riesgo = data.get('q_prgpro_riesgo', programacion.q_prgpro_riesgo)
            programacion.t_prgpro_estsit = data.get('t_prgpro_estsit', programacion.t_prgpro_estsit)
            programacion.t_prgpro_coment = data.get('t_prgpro_coment', programacion.t_prgpro_coment)
            programacion.estado_id = data.get('c_estado', programacion.estado_id)
            programacion.c_usuari_login = data.get('c_usuari_login', programacion.c_usuari_login)
            programacion.n_prgpro_pc = data.get('n_prgpro_pc', programacion.n_prgpro_pc)
            programacion.d_prgpro_fecdig = data.get('d_prgpro_fecdig', programacion.d_prgpro_fecdig)
            
            programacion.save()
            
            return {
                'success': True,
                'message': 'Programación actualizada exitosamente',
                'data': {
                    'c_prgpro': programacion.c_prgpro,
                    'c_proinv': programacion.proyecto.c_proinv,
                    'm_prgpro_mes': programacion.m_prgpro_mes,
                    'q_prgpro_financ': programacion.q_prgpro_financ,
                    'p_prgpro_fisica': programacion.p_prgpro_fisica,
                    'q_prgpro_caida': programacion.q_prgpro_caida,
                    'q_prgpro_increm': programacion.q_prgpro_increm,
                    'q_prgpro_riesgo': programacion.q_prgpro_riesgo,
                    't_prgpro_estsit': programacion.t_prgpro_estsit,
                    't_prgpro_coment': programacion.t_prgpro_coment,
                    'c_estado': programacion.estado.c_estado,
                    'c_usuari_login': programacion.c_usuari_login,
                    'n_prgpro_pc': programacion.n_prgpro_pc,
                    'd_prgpro_fecdig': programacion.d_prgpro_fecdig
                }
            }
            
    except ProgramacionProyectoInversion.DoesNotExist:
        return {
            'success': False,
            'message': 'La programación no existe',
            'data': None
        }
    except Exception as e:
        return {
            'success': False,
            'message': str(e),
            'data': None
        }

def format_sql_value(value):
    if value is None:
        return 'NULL'
    elif isinstance(value, str):
        return f"'{value}'"
    else:
        return value
    

def sql_to_pandas(sql):    
    server = os.environ.get("DB_HOST")
    database = "BDSIAF"
    username = os.environ.get("DB_USERNAME")
    password = os.environ.get("DB_PASSWORD")
    driver = 'ODBC Driver 17 for SQL Server'
    connection_string = f"mssql+pyodbc://{username}:{urllib.parse.quote_plus(password)}@{server}/{database}?driver={driver}"


    engine = create_engine(connection_string)
    

    df = pd.read_sql(sql, engine)

    # Cierra la conexión
    engine.dispose()

    return df

@api_view(['POST'])
@permission_classes([IsAuthenticated])        
def DownloadProyeccionMensualView(request):    
    ano_eje = request.data.get("ano_eje")
    mes_eje = request.data.get("mes_eje")
    sec_ejec = request.data.get("sec_ejec")
    
    if not all([ano_eje, mes_eje, sec_ejec]):
        return Response(
            data={"message": "Faltan parámetros"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        sincronizacion_id = obtener_ultima_sincronizacion_por_anio(ano_eje)

        if not sincronizacion_id:
            return Response(
                data={"message": "No se encontró la sincronización"},
                status=status.HTTP_404_NOT_FOUND
            )

        sql = f"""
        DECLARE @ANO_EJE decimal(18,2) = {format_sql_value(ano_eje)}
        DECLARE @MES_EJE decimal(18,2)= {format_sql_value(mes_eje)}
        DECLARE @SEC_EJEC char(10) = {format_sql_value(sec_ejec)}
        DECLARE @sincronizacion_id int = {format_sql_value(sincronizacion_id)}
        EXEC BDSIAF.DBO.SelectProyectosConGastosMensual @ANO_EJE=@ANO_EJE, @MES_EJE=@MES_EJE, @SEC_EJEC=@SEC_EJEC, @sincronizacion_id=@sincronizacion_id
        """ 
            
        df = sql_to_pandas(sql)  

        if len(df) > 0:
            # Calcular columnas solicitadas
            df['Proyección ajustada (d) = (a) - (b) + (c)'] = (
                df['q_prgpro_financ'].fillna(0) - df['q_prgpro_caida'].fillna(0) + df['q_prgpro_increm'].fillna(0)
            )
            # Calcular Avance financiero como decimal
            df['Avance financiero'] = (
                df['MONTO_DEVENGADO_MES'].fillna(0) / df['Proyección ajustada (d) = (a) - (b) + (c)']
            ).replace([float('inf'), -float('inf')], 0).fillna(0)

            # Redondear Avance financiero y Avance físico a cero decimales en porcentaje
            if 'Avance financiero' in df.columns:
                df['Avance financiero'] = (df['Avance financiero'] * 100).round(0) / 100
            if 'p_prgpro_fisica' in df.columns:
                df['p_prgpro_fisica'] = (df['p_prgpro_fisica'] / 100)
                df['p_prgpro_fisica'] = (df['p_prgpro_fisica'] * 100).round(0) / 100

            # Redondear columnas numéricas a 2 decimales (excepto Avance financiero y p_prgpro_fisica)
            columnas_redondear = [
                'q_prgpro_financ',
                'q_prgpro_caida',
                'q_prgpro_increm',
                'Proyección ajustada (d) = (a) - (b) + (c)',
                'MONTO_DEVENGADO_MES',
                'q_prgpro_riesgo'
            ]
            for col in columnas_redondear:
                if col in df.columns:
                    df[col] = df[col].round(2)

            # Seleccionar y renombrar columnas
            columnas = {
                'c_proinv_codigo': 'CUI',
                'n_proinv_nombre': 'Proyecto',
                'q_prgpro_financ': 'Programacion inicial (a)',
                'q_prgpro_caida': 'Caídas (b)',
                'q_prgpro_increm': 'Incrementos (c)',
                'Proyección ajustada (d) = (a) - (b) + (c)': 'Proyección ajustada (d) = (a) - (b) + (c)',
                'MONTO_DEVENGADO_MES': 'Ejecución (e)',
                'Avance financiero': 'Avance financiero',
                'p_prgpro_fisica': 'Avance físico',
                'q_prgpro_riesgo': 'Riesgo a fin de mes',
                't_prgpro_coment': 'Comentario'
            }
            df = df[list(columnas.keys())]
            df = df.rename(columns=columnas)

            # Guardar el DataFrame a Excel
            name_file = "ReporteGastosMensual"
            full_path_file = f"{PATH_TEMP}/{name_file}.xlsx"
            df.to_excel(full_path_file, index=False)

            # Abrir el archivo con openpyxl para aplicar formatos
            wb = openpyxl.load_workbook(full_path_file)
            ws = wb.active

            # Identificar las columnas por nombre
            col_indices = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

            # Columnas a formatear como moneda (separador de miles/millones)
            columnas_moneda = [
                'Programacion inicial (a)',
                'Caídas (b)',
                'Incrementos (c)',
                'Proyección ajustada (d) = (a) - (b) + (c)',
                'Ejecución (e)',
                'Riesgo a fin de mes'
            ]
            # Columnas a formatear como porcentaje
            columnas_porcentaje = [
                'Avance financiero',
                'Avance físico'
            ]

            # Aplicar formato a cada celda de las columnas correspondientes
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_name in columnas_moneda:
                    idx = col_indices.get(col_name)
                    if idx:
                        cell = row[idx-1]
                        cell.number_format = '#,##0.00'
                for col_name in columnas_porcentaje:
                    idx = col_indices.get(col_name)
                    if idx:
                        cell = row[idx-1]
                        cell.number_format = '0.00%'

            # Ajustar el ancho de las columnas automáticamente (autofit)
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Letra de la columna
                for cell in column_cells:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ''
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

            # Aplicar borde a toda la tabla (encabezados, datos y fila de sumatoria)
            thin_side = Side(style='thin', color="000000")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Agregar una fila de sumatoria al final (excepto Avance financiero y Avance físico)
            suma_dict = {}
            for col_name in ws[1]:
                header = col_name.value
                if header not in ['Avance financiero', 'Avance físico', 'Proyecto', 'Comentario', 'CUI']:
                    idx = col_name.column
                    col_letter = col_name.column_letter
                    # Sumar solo si la columna es numérica
                    try:
                        suma = sum([cell.value for cell in ws[col_letter][1:] if isinstance(cell.value, (int, float))])
                        suma_dict[header] = suma
                    except:
                        suma_dict[header] = ''
                else:
                    suma_dict[header] = ''
            # Crear la fila de sumatoria
            sum_row = []
            for col_name in ws[1]:
                header = col_name.value
                if header == 'Proyecto':
                    sum_row.append('TOTAL')
                elif header in suma_dict:
                    val = suma_dict[header]
                    # Si el valor es numérico, asegúrate de que sea float
                    if isinstance(val, (int, float)):
                        sum_row.append(float(val))
                    else:
                        sum_row.append('')
                else:
                    sum_row.append('')
            ws.append(sum_row)

            # Aplicar borde a la fila de sumatoria
            thin_side = Side(style='thin', color="000000")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.border = thin_border

            # Aplicar formato numérico y porcentaje a la fila de sumatoria (igual que a las filas de datos)
            last_row = ws.max_row
            for col_name in ws[1]:
                header = col_name.value
                idx = col_name.column
                cell = ws.cell(row=last_row, column=idx)
                if header in columnas_moneda:
                    cell.number_format = '#,##0.00'
                elif header in columnas_porcentaje:
                    cell.number_format = '0.00%'

            # Guardar el archivo con los formatos aplicados
            wb.save(full_path_file)

            with open(full_path_file, 'rb') as excel_file:
                response = HttpResponse(
                    excel_file.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            response['Content-Disposition'] = f'attachment; filename={name_file}.xlsx'
            os.remove(full_path_file)

            return response

        else:
            return Response(
                data={"message": "No hay datos para exportar"},
                status=status.HTTP_404_NOT_FOUND
            )

    except Exception as e:
        return Response(
            data={"message": str(e), "content": None},
            status=status.HTTP_400_BAD_REQUEST,
        )

def obtener_montos_por_ano(id_sincro: int):
    """
    Obtiene los montos PIA, PIM y DEVENGADO agrupados por año para una sincronización específica.
    
    Args:
        id_sincro (int): ID de la sincronización
        
    Returns:
        QuerySet: Conjunto de resultados con los montos agrupados por año
    """

    print(id_sincro)
    return RegistroSincronizacion.objects.filter(
        sincronizacion_id=id_sincro,
        tipo_act_proy=2
    ).values('ano_eje').annotate(
        MONTO_PIA=Sum('monto_pia'),
        MONTO_PIM=Sum('monto_pim'),
        MONTO_DEVENGADO=Sum('monto_devengado'),
        MES_EJE=Max('mes_eje')
    ).order_by('ano_eje')

class ObtenerMontosPorAnioView(RetrieveAPIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        anio =  request.query_params.get("anio")

        if not anio:
            return Response(
                data={"message": "Faltan parámetros", "content": {}},
                status=status.HTTP_400_BAD_REQUEST
            )

        id_sincro = obtener_ultima_sincronizacion_por_anio(anio)

        if not id_sincro:
            return Response(
                data={"message": "No se encontró la sincronización", "content": {}},
                status=status.HTTP_200_OK
            )
        montos = obtener_montos_por_ano(id_sincro)

        if montos:
            return Response(
                data={"message": "Montos obtenidos correctamente", "content": montos.first()},
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                data={"message": "No hay datos para exportar", "content": {}},
                status=status.HTTP_404_NOT_FOUND
            )

def contar_proyectos__por_anio(id_sincro: int, sec_ejec: str) -> int:
    """
    Cuenta el número de proyectos que tienen una suma de PIA y PIM mayor a 0.
    
    Args:
        id_sincro (int): ID de la sincronización
        sec_ejec (str): Código de la ejecutora
        
    Returns:
        int: Número de proyectos con montos positivos
    """
    
    return RegistroSincronizacion.objects.filter(
        sincronizacion_id=id_sincro,
        sec_ejec=sec_ejec,
        tipo_act_proy=2
    ).values('ano_eje', 'producto_proyecto').annotate(
        monto_presup=Sum('monto_pia') + Sum('monto_pim')
    ).filter(
        monto_presup__gt=0
    ).count()

class ContarProyectosPorAnioView(RetrieveAPIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        anio = request.query_params.get("anio")
        sec_ejec = request.query_params.get("sec_ejec")

        if not anio or not sec_ejec:
            return Response(
                data={"message": "Faltan parámetros", "content": {}},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        id_sincro = obtener_ultima_sincronizacion_por_anio(anio)

        if not id_sincro:
            return Response(
                data={"message": "No se encontró la sincronización", "content": {"total": 0}},
                status=status.HTTP_200_OK
            )

        return Response(
            data={"message": "Proyectos contados correctamente", "content": {"total": contar_proyectos__por_anio(id_sincro, sec_ejec)}},
            status=status.HTTP_200_OK
        )   


def predecir_ejecucion_futura(ejecucion_mes: list) -> list:
    """
    Predice la ejecución de meses futuros basándose en datos históricos usando estadísticas.
    
    Args:
        ejecucion_mes (list): Lista de diccionarios con la ejecución mensual histórica
            Cada diccionario debe tener: mes_eje, MONTO_DEVENGADO
            
    Returns:
        list: Lista de diccionarios con las predicciones mensuales
    """
    # Convertir los datos a arrays de numpy para cálculos más eficientes    
    meses = np.array([float(item['MesNumero']) for item in ejecucion_mes])    
    montos = np.array([float(item['MONTO_DEVENGADO']) for item in ejecucion_mes])    
    # Calcular la tendencia (pendiente de la línea de regresión)
    n = len(meses)
    if n < 2:
        return []
    
    # Calcular la pendiente (tendencia)
    x_mean = mean(meses)
    y_mean = mean(montos)
    numerador = sum((meses - x_mean) * (montos - y_mean))
    denominador = sum((meses - x_mean) ** 2)
    pendiente = numerador / denominador if denominador != 0 else 0    
    # Calcular la desviación estándar para el intervalo de confianza
    desviacion = float(stdev(montos)) if len(montos) > 1 else 0.0    
    # Calcular el promedio móvil de los últimos 3 meses
    ultimos_3_meses = montos[-3:] if len(montos) >= 3 else montos
    promedio_movil = float(mean(ultimos_3_meses))    
    # Nombres de los meses
    nombres_meses = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    # Generar predicciones para los meses restantes del año
    ultimo_mes = int(max(meses))
    predicciones = []    
    for mes in range(ultimo_mes + 1, 13):
        # Calcular la predicción base usando la tendencia
        prediccion_base = float(y_mean + pendiente * (mes - x_mean))
        # Ajustar la predicción usando el promedio móvil
        prediccion_ajustada = float((prediccion_base + promedio_movil) / 2)
        # Asegurar que la predicción no sea negativa
        prediccion_final = max(0.0, prediccion_ajustada)            
        # Calcular el intervalo de confianza
        intervalo_superior = float(prediccion_final + (1.96 * desviacion))
        intervalo_inferior = max(0.0, float(prediccion_final - (1.96 * desviacion)))        
        predicciones.append({
            'mes_eje': mes,
            'mes_nombre': nombres_meses[mes],
            'MONTO_DEVENGADO': round(prediccion_final, 2),
            'MONTO_MINIMO': round(intervalo_inferior, 2),
            'MONTO_MAXIMO': round(intervalo_superior, 2)
        })
    
    return predicciones

class EjecucionMesView(RetrieveAPIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        anio = request.query_params.get("anio")
        sec_ejec = request.query_params.get("sec_ejec")

        if not anio or not sec_ejec:
            return Response(
                data={"message": "Faltan parámetros", "content": {}},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        id_sincro = obtener_ultima_sincronizacion_por_anio(anio)

        if not id_sincro:
            return Response(
                data={"message": "No se encontró la sincronización", "content": []},
                status=status.HTTP_200_OK
            )
        
        ejecucion_mes = select_ejecucion_mes(id_sincro=id_sincro, sec_ejec=sec_ejec)

        if ejecucion_mes:
            return Response(
                data={"message": "Ejecución mes obtenida correctamente", "content": ejecucion_mes},
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                data={"message": "No hay datos para exportar", "content": []},
                status=status.HTTP_404_NOT_FOUND
            )

class EjecucionEsperadaView(RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        anio = request.query_params.get("anio")
        sec_ejec = request.query_params.get("sec_ejec")

        if not anio or not sec_ejec:
            return Response(
                data={"message": "Faltan parámetros", "content": {}},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        id_sincro = obtener_ultima_sincronizacion_por_anio(anio)

        if not id_sincro:
            return Response(
                data={"message": "No se encontró la sincronización", "content": {}},
                status=status.HTTP_200_OK
            )
        
        # Obtener datos históricos
        ejecucion_mes = select_ejecucion_mes(id_sincro=id_sincro, sec_ejec=sec_ejec)

        if ejecucion_mes:
            # Obtener predicciones
            predicciones = predecir_ejecucion_futura(ejecucion_mes)
            return Response(
                data={"message": "Ejecución esperada obtenida correctamente", "content": predicciones},
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                data={"message": "No hay datos para exportar", "content": {}},
                status=status.HTTP_404_NOT_FOUND
            )

def obtener_resumen_proyectos(id_sincro: int, sec_ejec: str):
    """
    Obtiene el resumen de proyectos por año, producto y montos.
    
    Args:
        id_sincro (int): ID de la sincronización
        sec_ejec (str): Código de la ejecutora
        
    Returns:
        QuerySet: Conjunto de resultados con el resumen de proyectos
    """
    from django.db.models import F, Sum, Q

    # Subconsulta para obtener los productos con suma de PIA + PIM > 0
    subquery = RegistroSincronizacion.objects.filter(
        sincronizacion_id=id_sincro,
        sec_ejec=sec_ejec,
        tipo_act_proy=2
    ).values(
        'producto_proyecto',
        'sincronizacion_id',
        'sec_ejec'
    ).annotate(
        total=Sum('monto_pia') + Sum('monto_pim')
    ).filter(
        total__gt=0
    )

    # Consulta principal usando INNER JOIN implícito
    return RegistroSincronizacion.objects.filter(
        sincronizacion_id=id_sincro,
        sec_ejec=sec_ejec,
        tipo_act_proy=2,
        producto_proyecto__in=subquery.values('producto_proyecto')
    ).values(
        'ano_eje',
        'producto_proyecto',
        'producto_proyecto_nombre'
    ).annotate(
        MONTO_PIA=Sum('monto_pia'),
        MONTO_PIM=Sum('monto_pim'),
        MONTO_DEVENGADO=Sum('monto_devengado')
    ).order_by('-MONTO_PIM')


class ResumenProyectosView(RetrieveAPIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request: Request):
        anio = request.query_params.get("anio")
        sec_ejec = request.query_params.get("sec_ejec")

        if not anio or not sec_ejec:
            return Response(
                data={"message": "Faltan parámetros", "content": {}},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        id_sincro = obtener_ultima_sincronizacion_por_anio(anio)

        if not id_sincro:
            return Response(
                data={"message": "No se encontró la sincronización", "content": []},
                status=status.HTTP_200_OK
            )
        
        resumen = obtener_resumen_proyectos(id_sincro, sec_ejec)
        
        if resumen: 
            return Response(
                data={"message": "Resumen de proyectos obtenido correctamente", "content": resumen},
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                data={"message": "No hay datos para exportar", "content": []},
                status=status.HTTP_200_OK
            )

@api_view(['POST'])
# @permission_classes([IsAuthenticated])        
def DownloadResumenProyectosView(request):    
    ano_eje = request.data.get("ano_eje")    
    sec_ejec = request.data.get("sec_ejec")
    
    if not all([ano_eje, sec_ejec]):
        return Response(
            data={"message": "Faltan parámetros"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        sincronizacion_id = obtener_ultima_sincronizacion_por_anio(ano_eje)

        if not sincronizacion_id:
            return Response(
                data={"message": "No se encontró la sincronización"},
                status=status.HTTP_404_NOT_FOUND
            )

        data = obtener_resumen_proyectos(sincronizacion_id, sec_ejec)
            
        df = pd.DataFrame(data)

        if len(df) > 0:
            # Convertir todas las columnas numéricas a float
            numeric_columns = ['MONTO_PIA', 'MONTO_PIM', 'MONTO_DEVENGADO']
            for col in numeric_columns:
                df[col] = df[col].astype(float)
            
            # Calcular Avance financiero como decimal, manejando división por cero
            df['Avance'] = df.apply(
                lambda row: (
                    row['MONTO_DEVENGADO'] / row['MONTO_PIM']
                    if row['MONTO_PIM'] != 0 else 0
                ),
                axis=1
            ).fillna(0)

            # Redondear Avance financiero a dos decimales en porcentaje
            if 'Avance' in df.columns:
                df['Avance'] = (df['Avance'] * 100).round(1) / 100

            if 'producto_proyecto' in df.columns:
                df['producto_proyecto'] = (df['producto_proyecto']).astype(int)
            
            # Redondear columnas numéricas a 2 decimales
            columnas_redondear = [
                'MONTO_PIA',
                'MONTO_PIM',
                'MONTO_DEVENGADO',                
            ]
            for col in columnas_redondear:
                if col in df.columns:
                    df[col] = df[col].round(2)

            # Seleccionar y renombrar columnas
            columnas = {                
                'producto_proyecto': 'CUI',
                'producto_proyecto_nombre': 'Descripción',
                'MONTO_PIA': 'PIA',
                'MONTO_PIM': 'PIM',
                'MONTO_DEVENGADO': 'Ejecución',
                'Avance': 'Avance',                
            }
            df = df[list(columnas.keys())]
            df = df.rename(columns=columnas)

            # Guardar el DataFrame a Excel
            name_file = "ReporteResumenProyectos"
            full_path_file = f"{PATH_TEMP}/{name_file}.xlsx"
            df.to_excel(full_path_file, index=False)

            # Abrir el archivo con openpyxl para aplicar formatos
            wb = openpyxl.load_workbook(full_path_file)
            ws = wb.active

            # Identificar las columnas por nombre
            col_indices = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

            # Columnas a formatear como moneda (separador de miles/millones)
            columnas_moneda = [
                'PIA',
                'PIM',
                'Ejecución',                
            ]
            # Columnas a formatear como porcentaje
            columnas_porcentaje = [
                'Avance',
            ]

            # Aplicar formato a cada celda de las columnas correspondientes
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_name in columnas_moneda:
                    idx = col_indices.get(col_name)
                    if idx:
                        cell = row[idx-1]
                        cell.number_format = '#,##0.00'
                for col_name in columnas_porcentaje:
                    idx = col_indices.get(col_name)
                    if idx:
                        cell = row[idx-1]
                        cell.number_format = '0.00%'

            # Ajustar el ancho de las columnas automáticamente (autofit)
            # for column_cells in ws.columns:
            #     max_length = 0
            #     column = column_cells[0].column_letter  # Letra de la columna
            #     for cell in column_cells:
            #         try:
            #             cell_value = str(cell.value) if cell.value is not None else ''
            #             if len(cell_value) > max_length:
            #                 max_length = len(cell_value)
            #         except:
            #             pass
            #     adjusted_width = max_length + 2
            #     ws.column_dimensions[column].width = adjusted_width

            # Aplicar borde a toda la tabla
            thin_side = Side(style='thin', color="000000")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Guardar el archivo con los formatos aplicados
            wb.save(full_path_file)

            with open(full_path_file, 'rb') as excel_file:
                response = HttpResponse(
                    excel_file.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            response['Content-Disposition'] = f'attachment; filename={name_file}.xlsx'
            os.remove(full_path_file)

            return response

        else:
            return Response(
                data={"message": "No hay datos para exportar"},
                status=status.HTTP_404_NOT_FOUND
            )

    except Exception as e:
        return Response(
            data={"message": str(e), "content": None},
            status=status.HTTP_400_BAD_REQUEST,
        )