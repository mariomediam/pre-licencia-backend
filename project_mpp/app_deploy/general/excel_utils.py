"""
Utilidades para exportación de datos a Excel

Este módulo proporciona funciones genéricas para exportar datos JSON a archivos Excel
con formato profesional.

Autor: Sistema
Fecha: 2026-03-02
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment


def create_excel_from_json(data, sheet_name="Datos", headers=None):
    """
    Crea un archivo Excel a partir de un JSON con formato profesional
    
    Args:
        data (list): Lista de diccionarios con los datos a exportar
        sheet_name (str, optional): Nombre de la hoja. Default: "Datos"
        headers (list, optional): Lista de headers personalizados. 
                                 Si no se proporciona, usa las keys del primer elemento
    
    Returns:
        io.BytesIO: Buffer con el archivo Excel generado
        
    Example:
        >>> datos = [
        ...     {"nombre": "Juan", "edad": 30},
        ...     {"nombre": "María", "edad": 25}
        ... ]
        >>> excel_buffer = create_excel_from_json(datos, "Personas")
        >>> # Retornar como respuesta HTTP
        >>> response = HttpResponse(excel_buffer.getvalue(), content_type="...")
    
    Features:
        - Headers con fondo azul (#4472C4) y texto blanco en negrita
        - Bordes delgados en todas las celdas
        - Ajuste automático del ancho de columnas (máximo 50 caracteres)
        - Formato automático de fechas ISO a dd/mm/yyyy HH:MM:SS
        - Centrado de headers horizontal y vertical
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if not data or len(data) == 0:
        # Si no hay datos, crear hoja vacía con headers si se proporcionaron
        if headers:
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
    else:
        # Obtener headers
        if headers is None:
            headers = list(data[0].keys())
        
        # Agregar headers
        sheet.append(headers)
        
        # Aplicar estilos a headers
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Agregar datos
        for row_data in data:
            row = []
            for header in headers:
                value = row_data.get(header, "")
                
                # Formatear fechas si es necesario
                if value and isinstance(value, str):
                    try:
                        # Intentar parsear fecha en formato ISO
                        if 'T' in value or '-' in value:
                            fecha_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
                            value = fecha_obj.strftime("%d/%m/%Y %H:%M:%S")
                    except:
                        pass
                
                row.append(value)
            
            sheet.append(row)
        
        # Aplicar bordes a todas las celdas con datos
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_excel_with_multiple_sheets(sheets_data):
    """
    Crea un archivo Excel con múltiples hojas
    
    Args:
        sheets_data (list): Lista de diccionarios con la configuración de cada hoja.
                           Cada diccionario debe tener:
                           - 'sheet_name' (str): Nombre de la hoja
                           - 'data' (list): Lista de diccionarios con los datos
                           - 'headers' (list, optional): Headers personalizados
    
    Returns:
        io.BytesIO: Buffer con el archivo Excel generado
        
    Example:
        >>> sheets = [
        ...     {
        ...         'sheet_name': 'Usuarios',
        ...         'data': [{"nombre": "Juan", "edad": 30}],
        ...         'headers': ['Nombre', 'Edad']
        ...     },
        ...     {
        ...         'sheet_name': 'Productos',
        ...         'data': [{"producto": "Laptop", "precio": 1200}]
        ...     }
        ... ]
        >>> excel_buffer = create_excel_with_multiple_sheets(sheets)
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remover la hoja por defecto
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for sheet_config in sheets_data:
        sheet_name = sheet_config.get('sheet_name', 'Datos')
        data = sheet_config.get('data', [])
        headers = sheet_config.get('headers')
        
        # Crear nueva hoja
        sheet = workbook.create_sheet(title=sheet_name)
        
        if not data or len(data) == 0:
            if headers:
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
        else:
            # Obtener headers
            if headers is None:
                headers = list(data[0].keys())
            
            # Agregar headers
            sheet.append(headers)
            
            # Aplicar estilos a headers
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Agregar datos
            for row_data in data:
                row = []
                for header in headers:
                    value = row_data.get(header, "")
                    
                    # Formatear fechas
                    if value and isinstance(value, str):
                        try:
                            if 'T' in value or '-' in value:
                                fecha_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                value = fecha_obj.strftime("%d/%m/%Y %H:%M:%S")
                        except:
                            pass
                    
                    row.append(value)
                
                sheet.append(row)
            
            # Aplicar bordes
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Ajustar ancho de columnas
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
